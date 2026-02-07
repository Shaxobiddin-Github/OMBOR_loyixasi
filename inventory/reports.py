import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from io import BytesIO
from django.db.models import Sum, F
from django.template.loader import get_template
from django.utils import timezone
from xhtml2pdf import pisa
from .models import Movement, Stock, Product

class ReportService:
    @staticmethod
    def generate_excel(data, headers, sheet_name="Report", title="Hisobot"):
        """
        Generates a professional Excel file.
        data: List of lists/tuples
        headers: List of strings
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Title Row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = center_align

        # Header Row
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data Rows
        for row_data in data:
            ws.append(row_data)
            # Apply border to current row
            current_row = ws.max_row
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.border = thin_border
                # Wrap text for longer fields if needed
                if isinstance(row_data[col_num-1], str) and len(row_data[col_num-1]) > 50:
                    cell.alignment = Alignment(wrap_text=True)

        # Auto-adjust column width (skip merged cells)
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip merged title)
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max(max_length + 2, 10)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_pdf(template_src, context):
        """
        Generates a PDF from a Django template.
        """
        template = get_template(template_src)
        html = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("utf-8")), result)
        if not pdf.err:
            result.seek(0)
            return result
        return None

    @staticmethod
    def get_stock_report_data():
        """
        Returns data for Stock Report (Excel/PDF context).
        """
        stocks = Stock.objects.select_related('product', 'product__category').order_by('product__category__name', 'product__name')
        
        # Excel Data
        excel_data = []
        for s in stocks:
            excel_data.append([
                s.product.sku,
                s.product.name,
                s.product.category.name,
                s.product.unit,
                s.current_qty,
                s.product.min_stock,
                "⚠️ KAM" if s.current_qty <= s.product.min_stock else "OK"
            ])
            
        # Context for PDF
        return {
            'excel_data': excel_data,
            'stocks': stocks,
            'total_items': stocks.count(),
            'total_qty': stocks.aggregate(Sum('current_qty'))['current_qty__sum'] or 0
        }

    @staticmethod
    def get_movement_report_data(start_date, end_date, movement_type=None):
        """
        Returns data for Movement Report.
        """
        movements = Movement.objects.filter(
            created_at__range=[start_date, end_date],
            status='VERIFIED'
        ).select_related('performed_by', 'face_employee').prefetch_related('items__product')

        if movement_type and movement_type != 'ALL':
            movements = movements.filter(movement_type=movement_type)

        excel_data = []
        for m in movements:
            items_str = ", ".join([f"{item.product.name} ({item.quantity})" for item in m.items.all()])
            excel_data.append([
                m.id,
                m.get_movement_type_display(),
                m.created_at.strftime("%Y-%m-%d %H:%M"),
                m.performed_by.username,
                m.face_employee.name if m.face_employee else "-",
                items_str,
                m.note
            ])

        return {
            'excel_data': excel_data,
            'movements': movements,
            'start_date': start_date,
            'end_date': end_date
        }
